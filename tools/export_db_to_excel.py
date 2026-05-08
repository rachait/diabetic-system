"""Export all SQLite tables + aggregated views to a single Excel file."""
import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH  = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'app.db'))
OUT_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'data', 'app_database_export_v2.xlsx'))

SKIP_TABLES = {'sqlite_sequence'}

HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
SECTION_FILL  = PatternFill("solid", fgColor="2E6DA4")
SECTION_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL      = PatternFill("solid", fgColor="EEF4FB")
EVEN_FILL     = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL    = PatternFill("solid", fgColor="E6F4EA")
RED_FILL      = PatternFill("solid", fgColor="FDECEA")
AMBER_FILL    = PatternFill("solid", fgColor="FFF8E1")
BORDER_SIDE   = Side(style='thin', color='C0C8D8')
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
BOLD_FONT     = Font(bold=True, size=10)

def safe_str(val):
    if val is None:
        return ""
    return str(val)

def write_header_row(ws, headers, row=1):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = CELL_BORDER

def write_data_cell(ws, row, col, value, fill=None, bold=False, align="left", num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill or (ALT_FILL if row % 2 == 0 else EVEN_FILL)
    c.border    = CELL_BORDER
    c.alignment = Alignment(horizontal=align, vertical="center")
    if bold:
        c.font = BOLD_FONT
    if num_format:
        c.number_format = num_format
    return c

def auto_width(ws, col_names, rows_data, start_col=1):
    for ci, col in enumerate(col_names, start=start_col):
        max_len = len(str(col))
        for row in rows_data:
            max_len = max(max_len, len(safe_str(row[ci - start_col])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row

tables = [r[0] for r in conn.execute(
    "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
).fetchall() if r[0] not in SKIP_TABLES]

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 1 – Summary (DB overview)
# ════════════════════════════════════════════════════════════════════════════════
ws_sum = wb.active
ws_sum.title = "Summary"

# Title banner
ws_sum.merge_cells("A1:D1")
title_cell = ws_sum["A1"]
title_cell.value     = f"Diabetes Prediction System – Database Export  ({datetime.now().strftime('%d %b %Y %H:%M')})"
title_cell.font      = Font(bold=True, color="FFFFFF", size=13)
title_cell.fill      = PatternFill("solid", fgColor="0D2B4E")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_sum.row_dimensions[1].height = 28

# DB info block
for r, (k, v) in enumerate([
    ("Database Type", "SQLite"),
    ("Path", DB_PATH),
    ("Exported At", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
], start=2):
    ws_sum.cell(row=r, column=1, value=k).font = BOLD_FONT
    ws_sum.cell(row=r, column=2, value=v)
    ws_sum.merge_cells(f"B{r}:D{r}")

# Table list header
write_header_row(ws_sum, ["Table Name", "Row Count", "Columns", "Notes"], row=6)

TABLE_NOTES = {
    "users":                 "Registered users (patients + admins)",
    "risk_assessments":      "Diabetes prediction results per user",
    "products":              "Health products available in store",
    "cart_items":            "Active shopping cart items",
    "orders":                "Customer orders",
    "order_items":           "Individual line items per order",
    "nutritionist_bookings": "Nutritionist session bookings",
    "meal_plans":            "Meal plans linked to products",
    "auth_activity_logs":    "Login / logout / security events",
    "doctors":               "Registered doctors for appointments",
    "appointments":          "Booked patient appointments",
}

for i, table in enumerate(tables, start=7):
    count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    cols  = [c[1] for c in conn.execute(f"PRAGMA table_info({table})").fetchall()]
    fill  = ALT_FILL if i % 2 == 0 else EVEN_FILL
    write_data_cell(ws_sum, i, 1, table, fill, bold=True)
    write_data_cell(ws_sum, i, 2, count, fill, align="center")
    write_data_cell(ws_sum, i, 3, ", ".join(cols), fill)
    write_data_cell(ws_sum, i, 4, TABLE_NOTES.get(table, ""), fill)

ws_sum.column_dimensions['A'].width = 28
ws_sum.column_dimensions['B'].width = 12
ws_sum.column_dimensions['C'].width = 70
ws_sum.column_dimensions['D'].width = 42

# ════════════════════════════════════════════════════════════════════════════════
# Sheet 2 – Users Overview (aggregated dashboard view)
# ════════════════════════════════════════════════════════════════════════════════
ws_usr = wb.create_sheet("Users Overview")

ws_usr.merge_cells("A1:K1")
bc = ws_usr["A1"]
bc.value     = "Users and Their Data"
bc.font      = Font(bold=True, color="FFFFFF", size=13)
bc.fill      = PatternFill("solid", fgColor="0D2B4E")
bc.alignment = Alignment(horizontal="center", vertical="center")
ws_usr.row_dimensions[1].height = 26

USER_COLS = [
    "User ID", "Username", "Full Name", "Email", "Role",
    "Total Assessments", "Diabetic Cases", "Latest Prediction",
    "Latest Confidence", "Latest Glucose", "Latest BMI",
    "Bookings", "Last Assessment Date", "Account Created"
]
write_header_row(ws_usr, USER_COLS, row=2)

user_rows = conn.execute(
    "SELECT id, username, full_name, email, is_admin, created_at FROM users ORDER BY id DESC"
).fetchall()

for ri, user in enumerate(user_rows, start=3):
    uid = user["id"]

    # Assessment stats
    assessments = conn.execute(
        "SELECT COUNT(*) FROM risk_assessments WHERE user_id=?", (uid,)
    ).fetchone()[0]

    diabetic = conn.execute(
        "SELECT COUNT(*) FROM risk_assessments WHERE user_id=? AND prediction LIKE '%Diabetic%' AND prediction NOT LIKE '%Non%'",
        (uid,)
    ).fetchone()[0]

    latest = conn.execute(
        "SELECT prediction, confidence, glucose, bmi, created_at FROM risk_assessments WHERE user_id=? ORDER BY created_at DESC LIMIT 1",
        (uid,)
    ).fetchone()

    bookings = conn.execute(
        "SELECT COUNT(*) FROM nutritionist_bookings WHERE user_id=?", (uid,)
    ).fetchone()[0]

    fill = ALT_FILL if ri % 2 == 0 else EVEN_FILL
    role = "Admin" if user["is_admin"] else "User"

    values = [
        uid,
        user["username"],
        user["full_name"] or "",
        user["email"],
        role,
        assessments,
        diabetic,
        latest["prediction"]  if latest else "N/A",
        f"{float(latest['confidence'])*100:.0f}%" if latest and latest["confidence"] not in (None, "") else "N/A",
        latest["glucose"]     if latest else "N/A",
        latest["bmi"]         if latest else "N/A",
        bookings,
        latest["created_at"]  if latest else "No records",
        user["created_at"] or "",
    ]

    for ci, val in enumerate(values, start=1):
        cell_fill = fill
        if ci == 5:   # Role column
            cell_fill = GREEN_FILL if role == "Admin" else EVEN_FILL
        if ci == 7 and isinstance(val, int) and val > 0:  # Diabetic cases
            cell_fill = AMBER_FILL
        write_data_cell(ws_usr, ri, ci, val, fill=cell_fill,
                        align="center" if ci in (1, 6, 7, 9, 10, 11, 12) else "left")

# Column widths for Users Overview
for col_letter, width in zip(
    [get_column_letter(i) for i in range(1, len(USER_COLS)+1)],
    [8, 18, 20, 30, 10, 18, 16, 30, 18, 15, 12, 10, 22, 22]
):
    ws_usr.column_dimensions[col_letter].width = width

ws_usr.merge_cells(f"A{len(user_rows)+3}:N{len(user_rows)+3}")
ws_usr.cell(row=len(user_rows)+3, column=1, value=f"Total users: {len(user_rows)}").font = BOLD_FONT

# ════════════════════════════════════════════════════════════════════════════════
# Remaining sheets – one per raw table
# ════════════════════════════════════════════════════════════════════════════════
for table in tables:
    rows      = conn.execute(f"SELECT * FROM {table}").fetchall()
    col_names = [c[1] for c in conn.execute(f"PRAGMA table_info({table})").fetchall()]

    ws = wb.create_sheet(title=table[:31])
    write_header_row(ws, col_names, row=1)

    for ri, row in enumerate(rows, start=2):
        fill = ALT_FILL if ri % 2 == 0 else EVEN_FILL
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=safe_str(val))
            cell.fill   = fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=False)

    for ci, col in enumerate(col_names, start=1):
        max_len = len(col)
        for row in rows:
            max_len = max(max_len, len(safe_str(row[ci - 1])))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

conn.close()
wb.save(OUT_PATH)
print(f"Exported {len(tables)} tables + Users Overview → {OUT_PATH}")
